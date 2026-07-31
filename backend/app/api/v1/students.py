from fastapi import APIRouter, Depends, status
from sqlalchemy.ext.asyncio import AsyncSession
from typing import Any, List

from app.schemas.student import StudentCreate, StudentResponse, StudentUpdate, StudentEnroll
from app.repositories.student import student_repo
from app.dependencies.database import get_db
from app.dependencies.auth import get_current_active_user, RequireRole
from app.models.user import User, Role
from app.models.student import Student
from passlib.context import CryptContext
from sqlalchemy import select

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

from app.core.exceptions import BadRequestException, NotFoundException, ForbiddenException
from app.services.student_service import StudentService

router = APIRouter()

@router.post("/", response_model=StudentResponse, status_code=status.HTTP_201_CREATED)
async def create_student(
    student_in: StudentCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(RequireRole(["admin"]))
) -> Any:
    """
    Create new student profile (Admin only).
    """
    if await student_repo.get_by_enrollment_number(db, enrollment_number=student_in.enrollment_number):
        raise BadRequestException("Student with this enrollment number already exists.")
    
    if await student_repo.get_by_user_id(db, user_id=student_in.user_id):
        raise BadRequestException("User already has a student profile.")

    return await student_repo.create(db, obj_in=student_in)

@router.post("/enroll", response_model=StudentResponse, status_code=status.HTTP_201_CREATED)
async def enroll_student(
    student_data: StudentEnroll,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(RequireRole(["admin"]))
) -> Any:
    """
    Enroll a new student by creating their User account and Student profile simultaneously.
    """
    return await StudentService.enroll_student(db, student_data)


from app.schemas.pagination import Pagination

@router.get("/", response_model=Pagination[StudentResponse])
async def read_students(
    skip: int = 0,
    limit: int = 100,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(RequireRole(["admin", "faculty"]))
) -> Any:
    """
    Retrieve all students (Admin/Faculty).
    """
    students = await student_repo.get_multi(db, skip=skip, limit=limit)
    # Get total count (for production, add a count() method to repo, doing len() for now)
    total = len(students) if students else 0
    pages = (total + limit - 1) // limit if limit > 0 else 1
    page = (skip // limit) + 1 if limit > 0 else 1
    
    return Pagination(
        items=students,
        total=total,
        page=page,
        size=limit,
        pages=pages
    )


@router.get("/me", response_model=StudentResponse)
async def read_student_me(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(RequireRole(["student"]))
) -> Any:
    """
    Get current student profile.
    """
    student = await student_repo.get_by_user_id(db, user_id=current_user.id)
    if not student:
        raise NotFoundException("Student profile not found")
    return student


@router.get("/{id}", response_model=StudentResponse)
async def read_student(
    id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
) -> Any:
    """
    Get student by ID.
    Students can only view their own profile. Admin/Faculty can view any.
    """
    student = await student_repo.get(db, id=id)
    if not student:
        raise NotFoundException("Student not found")
        
    if current_user.role.name == "student" and student.user_id != current_user.id:
        raise ForbiddenException()
        
    return student


@router.put("/{id}", response_model=StudentResponse)
async def update_student(
    id: int,
    student_in: StudentUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(RequireRole(["admin"]))
) -> Any:
    """
    Update student profile (Admin only).
    """
    student = await student_repo.get(db, id=id)
    if not student:
        raise NotFoundException("Student not found")
        
    update_data = student_in.model_dump(exclude_unset=True)
    full_name = update_data.pop("full_name", None)
    
    if full_name is not None:
        user_query = await db.execute(select(User).where(User.id == student.user_id))
        user = user_query.scalars().first()
        if user:
            user.full_name = full_name
            db.add(user)
            
    await student_repo.update(db, db_obj=student, obj_in=update_data)
    
    # Reload with joined user to avoid MissingGreenlet during serialization
    from sqlalchemy.orm import joinedload
    result = await db.execute(select(Student).options(joinedload(Student.user)).where(Student.id == id))
    return result.scalars().first()

@router.delete("/{id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_student(
    id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(RequireRole(["admin"]))
) -> None:
    """
    Delete student profile (Admin only).
    """
    success = await StudentService.delete_student(db, id)
    if not success:
        raise NotFoundException("Student not found")
    
    return None


# ── BULK IMPORT FROM CSV / EXCEL ──
from fastapi import UploadFile, File
import csv
import io

@router.post("/bulk-import")
async def bulk_import_students(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(RequireRole(["admin"]))
) -> Any:
    """
    Bulk import students from a CSV or Excel (.xlsx) file.
    Expected columns: full_name, email, password, enrollment_number, branch, semester, phone
    """
    filename = (file.filename or "").lower()
    content = await file.read()

    rows = []
    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append(row)
    elif filename.endswith(".xlsx") or filename.endswith(".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = [str(cell.value or "").strip().lower().replace(" ", "_") for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = str(val).strip() if val is not None else ""
            if any(row_dict.values()):
                rows.append(row_dict)
        wb.close()
    else:
        raise BadRequestException("Unsupported file format. Please upload a .csv or .xlsx file.")

    if not rows:
        raise BadRequestException("File is empty or has no data rows.")

    required_cols = {"full_name", "email", "password", "enrollment_number", "branch", "semester"}
    first_row_keys = {k.strip().lower().replace(" ", "_") for k in rows[0].keys()}
    missing = required_cols - first_row_keys
    if missing:
        raise BadRequestException(f"Missing required columns: {', '.join(missing)}. Required: full_name, email, password, enrollment_number, branch, semester, phone (optional)")

    success_count = 0
    errors = []

    for idx, row in enumerate(rows, start=2):
        try:
            full_name = row.get("full_name", "").strip()
            email = row.get("email", "").strip()
            password = row.get("password", "").strip()
            enrollment_number = row.get("enrollment_number", "").strip()
            branch = row.get("branch", "").strip()
            semester_str = row.get("semester", "1").strip()
            phone = row.get("phone", "").strip()

            if not full_name or not email or not password or not enrollment_number or not branch:
                errors.append({"row": idx, "error": "Missing required field(s)", "enrollment": enrollment_number or "N/A"})
                continue

            try:
                semester_val = int(float(semester_str))
            except (ValueError, TypeError):
                semester_val = 1

            student_data = StudentEnroll(
                full_name=full_name,
                email=email,
                password=password,
                enrollment_number=enrollment_number,
                branch=branch,
                semester=semester_val,
                phone=phone if phone else None
            )
            await StudentService.enroll_student(db, student_data)
            success_count += 1
        except BadRequestException as e:
            errors.append({"row": idx, "error": str(e.detail), "enrollment": row.get("enrollment_number", "N/A")})
        except Exception as e:
            errors.append({"row": idx, "error": str(e), "enrollment": row.get("enrollment_number", "N/A")})

    return {
        "message": f"Bulk import completed. {success_count} students enrolled successfully.",
        "success_count": success_count,
        "error_count": len(errors),
        "total_rows": len(rows),
        "errors": errors[:20]  # Return first 20 errors max
    }

